# views.py

from django.shortcuts import render, redirect
from .forms import ExcelUploadForm, InventarioUploadForm
from .models import Inven, OrdenDeVenta
from openpyxl import load_workbook
from .forms import OrdenDeVentaForm, ItemOfertaDeVentaForm


def borrar_orden(request, orden_id):
    orden = OrdenDeVenta.objects.get(id=orden_id)
    orden.delete()
    return redirect("odventa")


def procesar_inventario_excel(request):
    if request.method == "POST":
        form = InventarioUploadForm(request.POST, request.FILES)
        if form.is_valid():
            archivo_excel = request.FILES["archivo_excel"]

            # Cargar el archivo Excel con Openpyxl
            workbook = load_workbook(archivo_excel)
            sheet = workbook.active

            # Obtener datos de las celdas y guardar en la base de datos
            for row in sheet.iter_rows(min_row=2, values_only=True):
                Inven.objects.create(
                    nombre=row[0],
                    codigo=row[1],
                    precio=row[2],
                )

            return redirect("inventario")  # Redirige a la vista de inventarios
    else:
        form = InventarioUploadForm()

    return render(request, "formulario_inventario.html", {"form": form})


def procesar_ordenventa_excel(request):
    if request.method == "POST":
        form = ExcelUploadForm(
            request.POST, request.FILES
        )  # Asegúrate de utilizar la forma correcta
        if form.is_valid():
            archivo_excel = request.FILES["archivo_excel"]

            # Cargar el archivo Excel con Openpyxl
            workbook = load_workbook(archivo_excel)
            sheet = workbook.active

            # Obtener datos de las celdas y guardar en la base de datos
            for row in sheet.iter_rows(min_row=2, values_only=True):
                print("Datos del Excel:", row)

                OrdenDeVenta.objects.create(
                    codigo_sap=row[0],
                    nombre_proyecto=row[1],
                    direccion_proyecto=row[2],
                    fecha=row[3],
                    observacion=row[4],
                )

            return redirect(
                "odventa"
            )  # Ajusta la redirección según tu vista de orden de venta
    else:
        form = ExcelUploadForm()

    return render(request, "formulario_inventario.html", {"form": form})


def home(request):
    return render(request, "home.html")


def ofventa(request):
    return render(request, "ofventa.html")


def odventa(request):
    # Obtener todos los objetos OrdenDeVenta desde la base de datos
    ordenes_de_venta = OrdenDeVenta.objects.all()

    # Pasar los datos a la plantilla
    return render(request, "odventa.html", {"ordenes_de_venta": ordenes_de_venta})


def odcompra(request):
    return render(request, "odcompra.html")


def inventario(request):
    inventario_items = Inven.objects.all()
    return render(request, "inventario.html", {"inventario_items": inventario_items})


def crear_orden_venta(request):
    if request.method == "POST":
        orden_form = OrdenDeVentaForm(request.POST)
        item_form = ItemOfertaDeVentaForm(request.POST)

        if orden_form.is_valid() and item_form.is_valid():
            orden = orden_form.save()  # Guarda la orden

            # Asigna la orden al campo orden_venta antes de guardarlo
            item = item_form.save(commit=False)
            item.orden_venta = orden
            item.save()

            return redirect("crear_orden_venta")

    else:
        orden_form = OrdenDeVentaForm()
        item_form = ItemOfertaDeVentaForm()

    return render(
        request, "ofventa.html", {"orden_form": orden_form, "item_form": item_form}
    )
